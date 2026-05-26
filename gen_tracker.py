from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
import datetime

wb = Workbook()

# ── Palette ────────────────────────────────────────────────────
NAVY   = "1F3864"
BLUE   = "2E5FA3"
STEEL  = "4A90D9"
TEAL   = "0E7C86"
GOLD   = "C9A84C"
PURP   = "7B3F8E"
BRWN   = "B54A00"
WHITE  = "FFFFFF"
LGREY  = "F2F4F7"
MGREY  = "D6DCE4"
DGREY  = "595959"
LBLUE  = "EBF3FB"
STRIPE = "D6E4F5"
GREEN  = "1E7A43"
RED    = "C00000"

PILLAR_COLORS = [BLUE, TEAL, "C9762A", PURP, BRWN]
PILLAR_NAMES  = ["LLM & RLHF", "Inference HW", "Tech Sales", "Communication", "Voice"]
PILLAR_EMOJIS = ["🧠", "⚡", "🤝", "🗣️", "🎙️"]

# ── Style helpers ───────────────────────────────────────────────
def font(bold=False, size=11, color=None, name="Arial", italic=False):
    return Font(bold=bold, size=size, color=color or "000000", name=name, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(color="CCCCCC", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom(color=BLUE):
    t = Side(style="medium", color=color)
    n = Side(style=None)
    return Border(left=n, right=n, top=n, bottom=t)

def hdr(ws, row, col, text, bg=NAVY, fg=WHITE, sz=12, bold=True, align_h="center", wrap=False, merge_to=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=bold, size=sz, color=fg)
    c.fill = fill(bg)
    c.alignment = align(align_h, "center", wrap)
    c.border = border()
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return c

def cell(ws, row, col, value=None, bg=WHITE, fg="000000", sz=11, bold=False,
         align_h="left", wrap=False, num_fmt=None, border_color="CCCCCC"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=sz, color=fg)
    c.fill = fill(bg)
    c.alignment = align(align_h, "center", wrap)
    c.border = border(border_color)
    if num_fmt:
        c.number_format = num_fmt
    return c

def col_w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def row_h(ws, row, height):
    ws.row_dimensions[row].height = height


# ══════════════════════════════════════════════════════════════
#  TAB 1: DASHBOARD
# ══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📊 Dashboard"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

# Column widths
widths = [2, 28, 16, 16, 16, 16, 16, 16, 2]
for i, w in enumerate(widths, 1):
    col_w(ws, i, w)

# Title block
ws.merge_cells("B1:H1")
c = ws["B1"]
c.value = "🚀  PERSONAL GROWTH SPRINT 2026  —  Progress Dashboard"
c.font = Font(bold=True, size=18, color=WHITE, name="Arial")
c.fill = fill(NAVY)
c.alignment = align("left", "center")
row_h(ws, 1, 42)

ws.merge_cells("B2:H2")
c = ws["B2"]
c.value = "Saimir Baci  ·  June 1 – August 31, 2026  ·  5 Pillars  ·  2–3 hrs/day"
c.font = Font(italic=True, size=11, color="A8BAD0", name="Arial")
c.fill = fill(NAVY)
c.alignment = align("left", "center")
row_h(ws, 2, 22)

ws.merge_cells("B3:H3")
ws["B3"].fill = fill(NAVY)
row_h(ws, 3, 6)

row_h(ws, 4, 8)

# Section: Pillar summary
hdr(ws, 5, 2, "PILLAR", bg=NAVY, sz=11, align_h="left")
hdr(ws, 5, 3, "Target Hrs", bg=NAVY, sz=11)
hdr(ws, 5, 4, "Logged Hrs", bg=NAVY, sz=11)
hdr(ws, 5, 5, "% Complete", bg=NAVY, sz=11)
hdr(ws, 5, 6, "M1 Milestone", bg=NAVY, sz=11)
hdr(ws, 5, 7, "M2 Milestone", bg=NAVY, sz=11)
hdr(ws, 5, 8, "M3 Milestone", bg=NAVY, sz=11)
row_h(ws, 5, 22)

for i, (name, color, emoji) in enumerate(zip(PILLAR_NAMES, PILLAR_COLORS, PILLAR_EMOJIS)):
    r = 6 + i
    bg = LBLUE if i % 2 == 0 else WHITE
    cell(ws, r, 2, f"{emoji}  {name}", bg=bg, fg=NAVY, bold=True, sz=11)
    # Target hours: 45min LLM(3x/wk), 45min HW(3x/wk), 45min Sales(3x/wk), 30min Comm, 15min Voice
    targets = [72, 72, 72, 54, 36]
    cell(ws, r, 3, targets[i], bg=bg, fg=DGREY, sz=11, align_h="center")
    cell(ws, r, 4, f"='📅 Weekly Log'!{get_column_letter(3+i)}100", bg=bg, fg=BLUE, bold=True, sz=11, align_h="center")
    pct_col = get_column_letter(4)  # D
    cell(ws, r, 5, f"=IF(C{r}=0,0,D{r}/C{r})", bg=bg, fg=GREEN, bold=True, sz=11, align_h="center", num_fmt="0%")
    for mi in range(3):
        cell(ws, r, 6+mi, "⬜ Pending", bg=bg, fg=DGREY, sz=10, align_h="center")
    row_h(ws, r, 20)

# Legend for milestone status
r = 12
row_h(ws, r, 8)
r = 13
ws.merge_cells(f"B{r}:H{r}")
cell(ws, r, 2, "Milestone Status:   ⬜ Pending     🟡 In Progress     ✅ Complete     ❌ Missed", bg=LGREY, fg=DGREY, sz=10, wrap=True, align_h="left")
ws.merge_cells(f"B{r}:H{r}")
row_h(ws, r, 18)

# Section: Weekly streak
r = 15
ws.merge_cells(f"B{r}:H{r}")
cell(ws, r, 2, "WEEKLY HABIT STREAK  (update manually each Sunday)", bg=NAVY, fg=WHITE, bold=True, sz=12, align_h="left")
ws.merge_cells(f"B{r}:H{r}")
row_h(ws, r, 24)

r = 16
for wi in range(12):
    col = 2 + wi
    if col <= 8:
        hdr(ws, r, col, f"Wk {wi+1}", bg=STEEL, sz=10)
row_h(ws, r, 18)

r = 17
labels = ["Voice ✅?", "Study hrs", "Quality (1-5)"]
for li, label in enumerate(labels):
    rr = r + li
    cell(ws, rr, 2, label, bg=LBLUE, fg=NAVY, bold=True, sz=10)
    for wi in range(6):
        cell(ws, rr, 3+wi, "", bg=WHITE, fg=DGREY, sz=10, align_h="center")
    row_h(ws, rr, 18)

# Monthly reflection
r = 22
ws.merge_cells(f"B{r}:H{r}")
cell(ws, r, 2, "MONTHLY REFLECTIONS", bg=NAVY, fg=WHITE, bold=True, sz=12, align_h="left")
ws.merge_cells(f"B{r}:H{r}")
row_h(ws, r, 24)

months = ["June", "July", "August"]
month_clrs = [BLUE, TEAL, "C9762A"]
for mi, (mo, clr) in enumerate(zip(months, month_clrs)):
    r = 23 + mi * 5
    ws.merge_cells(f"B{r}:H{r}")
    cell(ws, r, 2, f"  {mo} 2026", bg=clr, fg=WHITE, bold=True, sz=11, align_h="left")
    ws.merge_cells(f"B{r}:H{r}")
    row_h(ws, r, 20)
    for qi, q in enumerate(["Top 3 wins:", "What to improve:", "Key insight:"]):
        rr = r + 1 + qi
        cell(ws, rr, 2, q, bg=LBLUE, fg=NAVY, bold=True, sz=10)
        ws.merge_cells(f"C{rr}:H{rr}")
        cell(ws, rr, 3, "", bg=WHITE, fg=DGREY, sz=10, wrap=True, align_h="left")
        row_h(ws, rr, 20)
    row_h(ws, r+4, 6)


# ══════════════════════════════════════════════════════════════
#  TAB 2: WEEKLY LOG
# ══════════════════════════════════════════════════════════════
wl = wb.create_sheet("📅 Weekly Log")
wl.sheet_view.showGridLines = False
wl.freeze_panes = "C5"

widths2 = [2, 14, 12, 10, 10, 10, 10, 10, 14, 16, 2]
for i, w in enumerate(widths2, 1):
    col_w(wl, i, w)

# Header
wl.merge_cells("B1:J1")
c = wl["B1"]
c.value = "📅  Daily Learning Log — June–August 2026"
c.font = Font(bold=True, size=16, color=WHITE, name="Arial")
c.fill = fill(NAVY)
c.alignment = align("left", "center")
row_h(wl, 1, 36)

wl.merge_cells("B2:J2")
c = wl["B2"]
c.value = "Log each session below. Complete the 'Note' column — even one sentence. Your streak matters more than your score."
c.font = Font(italic=True, size=10, color="A8BAD0", name="Arial")
c.fill = fill(NAVY)
c.alignment = align("left", "center")
row_h(wl, 2, 18)
row_h(wl, 3, 8)

# Column headers
headers = ["Date", "🧠 LLM\n(hrs)", "⚡ HW\n(hrs)", "🤝 Sales\n(hrs)", "🗣️ Comm\n(hrs)", "🎙️ Voice\n(Y/N)", "Total\n(hrs)", "Energy\n(1–5)", "Key Learning / Note"]
col_starts = [2, 3, 4, 5, 6, 7, 8, 9, 10]
col_clrs   = [NAVY, BLUE, TEAL, "C9762A", PURP, BRWN, NAVY, NAVY, NAVY]
for col, htext, clr in zip(col_starts, headers, col_clrs):
    h = wl.cell(row=4, column=col, value=htext)
    h.font = Font(bold=True, size=10, color=WHITE, name="Arial")
    h.fill = fill(clr)
    h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    h.border = border()
row_h(wl, 4, 32)

# Generate dates: June 1 – Aug 31, 2026 (skipping Sundays for study)
start = datetime.date(2026, 6, 1)
end   = datetime.date(2026, 8, 31)
dates = []
d = start
while d <= end:
    dates.append(d)
    d += datetime.timedelta(days=1)

for ri, d in enumerate(dates):
    row = 5 + ri
    is_sun = d.weekday() == 6
    bg = "FFF9F0" if is_sun else (LGREY if ri % 2 == 0 else WHITE)
    fg_date = "C9762A" if is_sun else NAVY

    # Date
    dc = wl.cell(row=row, column=2, value=d)
    dc.font = Font(bold=is_sun, size=10, color=fg_date, name="Arial")
    dc.fill = fill(bg)
    dc.alignment = align("center", "center")
    dc.border = border()
    dc.number_format = "DD MMM (DDD)"

    if is_sun:
        wl.cell(row=row, column=3, value="— Review Day —")
        wl.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
        c = wl.cell(row=row, column=3)
        c.font = Font(italic=True, size=10, color="C9762A", name="Arial")
        c.fill = fill(bg)
        c.alignment = align("center", "center")
        c.border = border()
        wl.cell(row=row, column=10, value="").fill = fill(bg)
    else:
        for ci, col in enumerate([3,4,5,6,7]):
            cc = wl.cell(row=row, column=col, value=None)
            cc.fill = fill(bg)
            cc.alignment = align("center", "center")
            cc.border = border()
            cc.number_format = "0.0"
        # Voice Y/N
        vc = wl.cell(row=row, column=7, value=None)
        vc.fill = fill(bg)
        vc.alignment = align("center")
        vc.border = border()
        # Total formula
        tot = wl.cell(row=row, column=8, value=f"=SUM(C{row}:F{row})")
        tot.font = Font(bold=True, size=10, color=NAVY, name="Arial")
        tot.fill = fill(bg)
        tot.alignment = align("center")
        tot.border = border()
        tot.number_format = "0.0"
        # Energy
        en = wl.cell(row=row, column=9, value=None)
        en.fill = fill(bg)
        en.alignment = align("center")
        en.border = border()
        # Note
        note = wl.cell(row=row, column=10, value=None)
        note.fill = fill(bg)
        note.alignment = align("left", "center", wrap=True)
        note.border = border()
    row_h(wl, row, 16)

# Totals row
total_row = 5 + len(dates)
wl.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True, size=11, color=WHITE, name="Arial")
wl.cell(row=total_row, column=2).fill = fill(NAVY)
wl.cell(row=total_row, column=2).alignment = align("center")
for ci, col in enumerate([3,4,5,6,8]):
    tc = wl.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{total_row-1})")
    tc.font = Font(bold=True, size=11, color=WHITE, name="Arial")
    tc.fill = fill(NAVY)
    tc.alignment = align("center")
    tc.number_format = "0.0"
wl.cell(row=total_row, column=7).fill = fill(NAVY)
wl.cell(row=total_row, column=9).fill = fill(NAVY)
wl.cell(row=total_row, column=10).fill = fill(NAVY)
row_h(wl, total_row, 22)


# ══════════════════════════════════════════════════════════════
#  HELPER: Create curriculum checklist tab
# ══════════════════════════════════════════════════════════════
def make_curriculum_tab(wb, tab_name, pillar_name, emoji, color, items_by_month):
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    widths = [2, 8, 40, 22, 12, 16, 2]
    for i, w in enumerate(widths, 1):
        col_w(ws, i, w)

    # Title
    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value = f"{emoji}  {pillar_name} — Curriculum & Progress"
    c.font = Font(bold=True, size=16, color=WHITE, name="Arial")
    c.fill = fill(color)
    c.alignment = align("left", "center")
    row_h(ws, 1, 36)
    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "Check off topics as you complete them. Add notes in the 'Notes' column as you go."
    c.font = Font(italic=True, size=10, color="D0D0D0", name="Arial")
    c.fill = fill(color)
    c.alignment = align("left", "center")
    row_h(ws, 2, 18)
    row_h(ws, 3, 8)

    # Col headers
    for col, label, clr in zip([2,3,4,5,6], ["Wk", "Topic / Exercise", "Resource", "Status", "Notes"], [color]*5):
        h = ws.cell(row=4, column=col, value=label)
        h.font = Font(bold=True, size=11, color=WHITE, name="Arial")
        h.fill = fill(clr)
        h.alignment = align("center" if col != 4 else "left", "center")
        h.border = border()
    row_h(ws, 4, 22)

    r = 5
    for month_label, items in items_by_month:
        # Month separator
        ws.merge_cells(f"B{r}:F{r}")
        mc = ws.cell(row=r, column=2, value=f"  {month_label}")
        mc.font = Font(bold=True, size=12, color=WHITE, name="Arial")
        mc.fill = fill(color)
        mc.alignment = align("left", "center")
        mc.border = border()
        row_h(ws, r, 22)
        r += 1

        for wk, topic, resource in items:
            bg = LGREY if r % 2 == 0 else WHITE
            cell(ws, r, 2, wk, bg=bg, fg=DGREY, sz=10, align_h="center")
            cell(ws, r, 3, topic, bg=bg, fg="1A1A2E", sz=10, wrap=True, align_h="left")
            cell(ws, r, 4, resource, bg=bg, fg=DGREY, sz=10, wrap=True, align_h="left")
            st = ws.cell(row=r, column=5, value="⬜ Todo")
            st.font = Font(size=10, color=DGREY, name="Arial")
            st.fill = fill(bg)
            st.alignment = align("center", "center")
            st.border = border()
            cell(ws, r, 6, "", bg=bg, fg=DGREY, sz=10, wrap=True, align_h="left")
            row_h(ws, r, 30)
            r += 1

        row_h(ws, r, 8)
        r += 1

    return ws


# ── LLM Tab ────────────────────────────────────────────────────
make_curriculum_tab(wb, "🧠 LLM & RLHF", "LLM Architectures & RLHF", "🧠", BLUE, [
    ("MONTH 1: JUNE — Foundations", [
        ("W1", "Transformer architecture: multi-head attention, positional encoding, layer norm", "Annotated Transformer (Harvard NLP); Karpathy 'makemore' / 'nanoGPT'"),
        ("W1", "Self-attention mechanics: Q, K, V matrices; scaled dot-product attention", "Attention Is All You Need (Vaswani et al., 2017)"),
        ("W2", "Tokenisation: BPE, SentencePiece, byte-level BPE; build a tiny tokeniser", "Karpathy 'minBPE' repo"),
        ("W2", "Pre-training objective: causal language modelling, next-token prediction, loss landscape", "Scaling Laws paper (Kaplan et al., 2020)"),
        ("W3", "Supervised Fine-Tuning (SFT): datasets, instruction formats, loss masking", "Hugging Face TRL docs; Alpaca paper"),
        ("W3", "Parameter-efficient fine-tuning: LoRA, QLoRA, prefix tuning — implement LoRA layer", "LoRA paper (Hu et al., 2021); LlamaFactory repo"),
        ("W4", "RLHF: reward modelling from preferences, PPO algorithm basics", "InstructGPT paper (Ouyang et al., 2022)"),
        ("W4", "PPO implementation: policy, value head, KL penalty, clip objective", "CleanRL PPO implementation; TRL PPO trainer"),
    ]),
    ("MONTH 2: JULY — Advanced Alignment", [
        ("W5", "Direct Preference Optimisation (DPO): derivation, why it avoids RL instability", "DPO paper (Rafailov et al., 2023)"),
        ("W5", "RLAIF and Constitutional AI: self-critique loop, red-teaming with Claude", "Constitutional AI paper (Bai et al., 2022)"),
        ("W6", "Mixture of Experts (MoE): routing, sparse activation, Mixtral architecture", "Switch Transformer paper; Mixtral technical report"),
        ("W6", "State-space models: S4, Mamba, selective state spaces — conceptual understanding", "Mamba paper (Gu & Dao, 2023)"),
        ("W7", "LLM evaluation: MMLU, HellaSwag, TruthfulQA; benchmark pitfalls and Goodhart's Law", "Big-Bench; HELM evaluation framework docs"),
        ("W7", "Safety evaluation: red-teaming techniques, jailbreak categories, refusal training", "Anthropic red-teaming paper; Llama Guard"),
        ("W8", "RAG architecture: chunking, embedding, vector search (FAISS/Chroma), reranking", "LangChain / LlamaIndex docs; ColBERT paper"),
        ("W8", "Hybrid search: BM25 + dense retrieval; reciprocal rank fusion", "BEIR benchmark paper"),
    ]),
    ("MONTH 3: AUGUST — Applied & Frontier", [
        ("W9",  "Multi-modal models: ViT image encoder, LLaVA architecture, image-text alignment", "LLaVA paper (Liu et al., 2023); Flamingo paper"),
        ("W9",  "Vision tokenisation: patch embeddings, cross-attention vs early fusion", "BLIP-2 paper"),
        ("W10", "LLM agents: ReAct prompting, tool-use, function calling, memory types", "ReAct paper; OpenAI function calling docs"),
        ("W10", "Agentic loops: planning, reflection, multi-agent coordination (brief survey)", "AutoGPT, CrewAI, LangGraph docs"),
        ("W11", "Quantisation: GPTQ, GGUF, AWQ — what changes in the weights, what doesn't", "GPTQ paper; llama.cpp repo"),
        ("W11", "Speculative decoding: draft model, verification, throughput gains", "Speculative decoding paper (Leviathan et al., 2022)"),
        ("W12", "Personal synthesis: write 2-page technical brief on one frontier topic", "Your choice — publish on LinkedIn or internal doc"),
        ("W12", "Review: re-read your Week 1 notes; document how your mental model changed", "Your own notes from Weeks 1–11"),
    ]),
])

# ── Hardware Tab ───────────────────────────────────────────────
make_curriculum_tab(wb, "⚡ Inference HW", "Inference Hardware Architecture", "⚡", TEAL, [
    ("MONTH 1: JUNE — GPU Fundamentals", [
        ("W1", "GPU architecture: SM hierarchy, CUDA cores, warp scheduling, occupancy", "NVIDIA CUDA Programming Guide Ch.1–4"),
        ("W1", "Memory hierarchy: registers → shared memory → L2 → DRAM; bandwidth vs latency", "Tim Dettmers 'GPU for Deep Learning' blog"),
        ("W2", "CUDA kernel writing: thread blocks, grids, shared memory tile; profile with Nsight", "CUDA by Example (book); GPU Mode Lecture 1"),
        ("W2", "Matrix multiply: naive → tiled → cuBLAS; measure FLOP/s and bandwidth utilisation", "CUTLASS repo; Simon Boehm blog on GEMM"),
        ("W3", "Inference bottlenecks: arithmetic intensity, memory-bound vs compute-bound, roofline model", "Roofline model paper; Dissecting the A100 (MLSys 2022)"),
        ("W3", "KV cache: why it exists, memory footprint formula, GQA / MQA variants", "Efficient Transformers survey; GQA paper"),
        ("W4", "FlashAttention: IO-aware tiling, SRAM reuse, backward pass summary", "FlashAttention-2 paper (Dao, 2023)"),
        ("W4", "Paged attention: vLLM's PagedAttention — OS memory management analogy", "vLLM paper (Kwon et al., 2023)"),
    ]),
    ("MONTH 2: JULY — Metal, macOS & AMD", [
        ("W5", "Apple Silicon: Unified Memory architecture, bandwidth, Neural Engine TOPS", "Apple Silicon M4 whitepaper; MLX docs"),
        ("W5", "Metal compute shaders: thread groups, threadgroup memory — conceptual vs CUDA", "Apple Metal Shading Language Spec"),
        ("W6", "MLX framework: run Mistral-7B / Phi-3 on Mac; measure tokens/sec and memory use", "MLX GitHub repo; mlx-lm examples"),
        ("W6", "MLX internals: lazy evaluation, unified memory ops, metal backend", "MLX design doc; Awni Hannun talk"),
        ("W7", "AMD RDNA/CDNA architecture: wavefronts, LDS, MI300X for inference", "AMD ROCm docs; AMD MI300X whitepaper"),
        ("W7", "HIP programming: porting a CUDA kernel to HIP; ROCm profiling tools", "ROCm HIP docs; Hipify tool"),
        ("W8", "Operator fusion: fusing LayerNorm + attention; why it reduces memory traffic", "Triton programming guide; flash-attention Triton kernel"),
        ("W8", "INT8 / INT4 quantised kernels: GPTQ kernel, AWQ kernel — how weights are packed", "GPTQ kernel repo; AWQ paper"),
    ]),
    ("MONTH 3: AUGUST — Custom Silicon & Serving", [
        ("W9",  "Groq LPU: deterministic SRAM-only execution, streaming, compiler-driven parallelism", "Groq architecture whitepaper; Groq blog posts"),
        ("W9",  "Inference latency vs throughput trade-offs on Groq vs GPU", "Groq benchmark comparisons"),
        ("W10", "Cerebras wafer-scale: dataflow architecture, no memory hierarchy, sparse tensor cores", "Cerebras CS-3 whitepaper"),
        ("W10", "Qualcomm AI 100 Ultra: heterogeneous compute, on-device inference for edge robotics", "Qualcomm AI 100 whitepaper; QNN SDK docs"),
        ("W11", "vLLM serving: continuous batching, PagedAttention, multi-GPU tensor parallelism", "vLLM GitHub; Anyscale serving blog"),
        ("W11", "TensorRT-LLM: NVIDIA's high-performance inference engine, plugin system", "TensorRT-LLM docs; NVIDIA blog"),
        ("W12", "Personal synthesis: hardware selection guide for an edge robotics inference workload", "Combine roofline, chip specs, and power budget"),
        ("W12", "Compare: latency/power/cost table for GPU vs Apple M4 vs Groq for a given model", "Benchmarks you ran + published data"),
    ]),
])

# ── Sales Tab ─────────────────────────────────────────────────
make_curriculum_tab(wb, "🤝 Tech Sales", "Technical Sales (Robotics)", "🤝", "C9762A", [
    ("MONTH 1: JUNE — Methodology Foundations", [
        ("W1", "Read 'The Challenger Sale': core Challenger model, teaching, tailoring, taking control", "The Challenger Sale — Dixon & Adamson"),
        ("W1", "MEDDIC: Metrics, Economic Buyer, Decision Criteria, Decision Process, Identify Pain, Champion", "MEDDIC Academy (meddic.com)"),
        ("W2", "ICP definition: map logistics AMR / agriculture / manufacturing cobot customer profiles", "Augmentifai Sales GTM Skill (internal)"),
        ("W2", "Buyer personas: Fleet Ops Manager, CTO, DevOps Lead, SRE — pain, language, objections", "Augmentifai Sales GTM Skill (internal)"),
        ("W3", "Discovery call framework: 5 high-value opening questions; active listening drills", "SPIN Selling — Neil Rackham (Ch. 1–4)"),
        ("W3", "Economic Buyer identification: signs you're talking to a champion vs decision-maker", "MEDDIC Academy"),
        ("W4", "Value proposition: craft 30-sec / 2-min / 10-min pitch for Synapse Debugger", "Augmentifai concept + company skills (internal)"),
        ("W4", "Record and review your pitch: watch for pace, filler, and clarity of the 'so what'", "Voice Memos / Loom"),
    ]),
    ("MONTH 2: JULY — Robotics Domain & Demo", [
        ("W5", "ROS2 fundamentals: nodes, topics, services, actions; common robotics failure modes", "ROS2 Documentation (docs.ros.org); YouTube intro"),
        ("W5", "Fleet management systems: FMS architecture, common vendors, pain points at scale", "Vendor websites; Robotics Business Review"),
        ("W6", "Competitive battle card: top 3 Augmentifai competitors — differentiators, weaknesses", "Augmentifai Sales GTM Skill (internal)"),
        ("W6", "Objection handling: practise 5 common objections aloud with responses", "Gong Revenue Intelligence Blog"),
        ("W7", "Demo structure: problem → demo → value — 20-minute Synapse Debugger demo outline", "Augmentifai Synapse Skill (internal)"),
        ("W7", "Record and review the demo: note where energy drops and where clarity is low", "Loom / Voice Memos"),
        ("W8", "Proof-of-value (POV) design: success criteria, timeline, owner, metrics, exit criteria", "MEDDIC POV framework"),
        ("W8", "POV objection: 'we don't have time' — practice the response 5 times", "30 Min to President's Club Podcast"),
    ]),
    ("MONTH 3: AUGUST — Pipeline & Applied Selling", [
        ("W9",  "Write 10 personalised cold outreach emails targeting logistics AMR ICP", "Smart Brevity framework; Augmentifai GTM Skill"),
        ("W9",  "LinkedIn outreach: 5 targeted connection messages with a clear value hook", "LinkedIn Sales Navigator best practices"),
        ("W10", "Negotiation basics: BATNA, anchoring, concession ladder, 'if you... then I...'", "Never Split the Difference — Chris Voss (Ch. 1–3)"),
        ("W10", "Role-play 2 closing scenarios: budget objection and 'need to think about it'", "Practice partner or record solo"),
        ("W11", "CRM pipeline discipline: build mock pipeline in Friday PM with 5 dummy deals", "Friday PM MCP tools (internal)"),
        ("W11", "Weekly deal review: practise 3-min verbal update per deal — status, next step, risk", "Salesforce / Gong deal review methodology"),
        ("W12", "Full mock sales cycle: discovery → demo → objection → close (full recorded session)", "All of the above"),
        ("W12", "Debrief and score: use MEDDIC scorecard on the mock deal — what would you do differently?", "MEDDIC Academy scorecard template"),
    ]),
])

# ── Communication Tab ─────────────────────────────────────────
make_curriculum_tab(wb, "🗣️ Communication", "Communication Skills", "🗣️", PURP, [
    ("MONTH 1: JUNE — Structure & Clarity", [
        ("W1", "Read The Pyramid Principle: SCQA framework, inductive vs deductive ordering", "The Pyramid Principle — Barbara Minto (Ch. 1–3)"),
        ("W1", "Apply SCQA: rewrite 3 real emails using the Situation-Complication-Question-Answer structure", "Your own inbox"),
        ("W2", "Writing audit: pick 10 recent messages — highlight vague words, buried main points, padding", "Hemingway App (hemingwayapp.com)"),
        ("W2", "Find your filler patterns: list your top 5 vague/weak words (basically, very, sort of, just)", "Your own emails and Slack messages"),
        ("W3", "One-idea-per-sentence: rewrite 5 paragraphs; target max 20 words per sentence", "On Writing Well — Zinsser (Ch. 2–4)"),
        ("W3", "Lead with the conclusion: practice starting every response with the answer, not the context", "Smart Brevity — Axios"),
        ("W4", "1-sentence summary discipline: after every meeting, write one sentence summarising the decision", "Apply in Friday PM task notes"),
        ("W4", "Weekly writing challenge: write a 5-bullet weekly update that a C-level could read in 60 seconds", "BLUF format guide"),
    ]),
    ("MONTH 2: JULY — Conciseness & Precision", [
        ("W5", "Active voice practice: identify passive sentences in 5 docs/emails; rewrite all", "Elements of Style — Strunk & White (Rule 14)"),
        ("W5", "Cut 20%: take any piece of writing and reduce it by 20% without losing meaning", "On Writing Well — Ch. 7"),
        ("W6", "Verbal 3-point answers: practice the 'position → reason → example' structure in conversations", "Toastmasters Table Topics format"),
        ("W6", "Record 3 verbal responses to 'tell me about X' questions; listen for filler and rambling", "Voice Memos app"),
        ("W7", "The smart brief: write a 5-line project update that replaces a 30-minute meeting", "Axios Smart Brevity format"),
        ("W7", "Async-first writing: practice the principle of 'no meeting if a message would do'", "Doist's Guide to Async Communication"),
        ("W8", "Presentation clarity: for any 10-min topic, write an opening hook (1 sentence), 3 points, close", "TED Talk structure analysis"),
        ("W8", "Peer feedback: send 3 emails to a trusted colleague and ask for honest clarity feedback", "Use a simple scorecard: clear? concise? actionable?"),
    ]),
    ("MONTH 3: AUGUST — Executive Communication", [
        ("W9",  "BLUF for leadership: practice 'Bottom Line Up Front' format for any upward communication", "US Army BLUF doctrine; exec writing guides"),
        ("W9",  "Executive email template: subject (outcome), 2-line summary, 3 bullets, clear ask", "Your own templates"),
        ("W10", "Presentation structure: opening hook → insight 1 → 2 → 3 → call to action. 10-min practice talk", "Garr Reynolds 'Presentation Zen' (book/blog)"),
        ("W10", "Speaking clarity drill: record yourself explaining a complex topic to a 12-year-old", "Voice Memos; share with a non-technical friend"),
        ("W11", "Feedback loop: ask a colleague weekly for one thing to improve in your communication", "Use a specific question: 'Was my last update clear and direct?'"),
        ("W11", "Conflict communication: practice delivering difficult news clearly and kindly (role-play)", "Crucial Conversations — Patterson et al. (Ch. 1–3)"),
        ("W12", "Personal style guide: document your 10 communication rules — your own cheat sheet", "Print and keep visible at desk"),
        ("W12", "Before/after portfolio: collect 5 examples of improved writing from Week 1 to Week 12", "Side-by-side comparison document"),
    ]),
])

# ── Voice Tab ─────────────────────────────────────────────────
make_curriculum_tab(wb, "🎙️ Voice", "Voice & Command", "🎙️", BRWN, [
    ("MONTH 1: JUNE — Breath & Resonance (15–20 min/day)", [
        ("Daily", "Diaphragmatic breathing: 4-2-6 count, belly-first, 3 minutes", "Roger Love 'Set Your Voice Free' (Ch. 1–2)"),
        ("Daily", "Humming + lip trills: start at comfortable pitch, slide up and down, 3 minutes", "Roger Love vocal warmup exercises"),
        ("Daily", "Articulation drills: 'Red leather yellow leather' / 'Unique New York' at slow then fast pace", "Toastmasters tongue twister sets"),
        ("Daily", "Paced reading aloud: any text, 80% of natural speed, 4 minutes; record on phone", "Your own voice memos"),
        ("Weekly", "Listen back to Monday's recording; note one thing to improve", "Voice Memos app"),
        ("W1–2", "Posture awareness: stand or sit tall when speaking; check shoulder and jaw tension", "Julian Treasure TED talk on speaking"),
        ("W3–4", "Resonance placement: feel the vibration in chest vs head voice; practise both registers", "Roger Love 'Set Your Voice Free' (Ch. 3)"),
        ("W4",   "Baseline recording: record a 2-minute pitch. This is your Month 1 reference.", "Loom or Voice Memos — label clearly"),
    ]),
    ("MONTH 2: JULY — Pace, Articulation & Pause", [
        ("Daily", "Continue daily 16-min routine from Month 1", "Habit maintained from June"),
        ("Daily", "Deliberate pause practice: read a paragraph aloud and insert a 1-second pause after each sentence", "Any text; focus on resisting the urge to rush"),
        ("Daily", "Transcribe 1 minute of your own recording using Otter.ai; count filler words", "Otter.ai (free tier)"),
        ("W5",   "Consonant crispness: over-articulate 't', 'd', 'p', 'b' at the end of words — then normalise", "Classic actor vocal training exercises"),
        ("W6",   "Pace variation: read the same paragraph at 3 speeds (slow, normal, fast) — feel the difference", "Your own recordings; compare the effect"),
        ("W7",   "Record a 3-min monologue on any technical topic; count: umms, likes, you-knows", "Voice Memos; review with transcript"),
        ("W7",   "Eliminate top filler word for 1 week: replace 'um' with a pause — measure the change", "Transcript comparison"),
        ("W8",   "Present a 5-min topic to a friend or colleague; ask for pace and clarity feedback", "Real conversation practice"),
    ]),
    ("MONTH 3: AUGUST — Projection, Tonality & Command", [
        ("Daily", "Continue full daily routine; add 2 min of projection practice (speak to back of room)", "Project from diaphragm, not throat"),
        ("Daily", "Vocal variety drill: read the same sentence with 3 different emotions — engage your tonal range", "Any paragraph; experiment freely"),
        ("W9",   "Confident word endings: do NOT drop your voice at the end of statements (downward inflection)", "Record and check: does your energy stay up to the last word?"),
        ("W10",  "Tonality map: identify your 3 go-to tones (explaining, persuading, questioning) — practise each", "Roger Love on tonal variety"),
        ("W11",  "Live presentation: deliver a 5-min talk in a real meeting, webinar, or event", "Request 2 pieces of specific voice feedback afterwards"),
        ("W11",  "Synapse Debugger pitch: record the full 2-min pitch — your Month 3 reference recording", "Compare to your Month 1 baseline from Week 4"),
        ("W12",  "Side-by-side comparison: listen to Month 1 vs Month 3 recordings; write what changed", "Note: pace, resonance, filler frequency, confidence"),
        ("W12",  "Voice maintenance plan: write a 10-min weekly routine you will continue after August", "Your own reference document"),
    ]),
])

# ══════════════════════════════════════════════════════════════
#  TAB: RESOURCES
# ══════════════════════════════════════════════════════════════
rws = wb.create_sheet("📚 Resources")
rws.sheet_view.showGridLines = False
widths_r = [2, 36, 14, 14, 22, 2]
for i, w in enumerate(widths_r, 1):
    col_w(rws, i, w)

rws.merge_cells("B1:E1")
c = rws["B1"]
c.value = "📚  Master Resource Library"
c.font = Font(bold=True, size=16, color=WHITE, name="Arial")
c.fill = fill(NAVY)
c.alignment = align("left", "center")
row_h(rws, 1, 36)
rws.merge_cells("B2:E2")
c = rws["B2"]
c.value = "All resources for the 3-month plan. Use the 'Status' column to track: Not Started / Reading / Done."
c.font = Font(italic=True, size=10, color="A8BAD0", name="Arial")
c.fill = fill(NAVY)
c.alignment = align("left", "center")
row_h(rws, 2, 18)
row_h(rws, 3, 8)

for col, label, clr in zip([2,3,4,5], ["Title / Resource", "Type", "Pillar", "Status"], [NAVY]*4):
    h = rws.cell(row=4, column=col, value=label)
    h.font = Font(bold=True, size=11, color=WHITE, name="Arial")
    h.fill = fill(clr)
    h.alignment = align("left" if col==2 else "center", "center")
    h.border = border()
row_h(rws, 4, 22)

all_resources = [
    ("Andrej Karpathy — 'Neural Networks: Zero to Hero'", "YouTube", "🧠 LLM"),
    ("Attention Is All You Need (Vaswani et al., 2017)", "Paper", "🧠 LLM"),
    ("InstructGPT Paper (Ouyang et al., 2022)", "Paper", "🧠 LLM"),
    ("DPO Paper (Rafailov et al., 2023)", "Paper", "🧠 LLM"),
    ("Constitutional AI Paper (Bai et al., 2022)", "Paper", "🧠 LLM"),
    ("Lilian Weng's Blog (lilianweng.github.io)", "Blog", "🧠 LLM"),
    ("Sebastian Raschka — 'Build an LLM from Scratch'", "Book/Code", "🧠 LLM"),
    ("Hugging Face Course", "Course", "🧠 LLM"),
    ("Mamba Paper (Gu & Dao, 2023)", "Paper", "🧠 LLM"),
    ("NVIDIA CUDA Programming Guide", "Docs", "⚡ HW"),
    ("Tim Dettmers — 'GPU for Deep Learning' blog", "Blog", "⚡ HW"),
    ("FlashAttention-2 Paper (Dao, 2023)", "Paper", "⚡ HW"),
    ("vLLM Paper (Kwon et al., 2023)", "Paper", "⚡ HW"),
    ("Apple MLX Documentation", "Docs", "⚡ HW"),
    ("Groq Architecture Whitepaper", "Whitepaper", "⚡ HW"),
    ("GPU Mode Lecture Series", "Video", "⚡ HW"),
    ("Roofline Model Paper (Williams et al.)", "Paper", "⚡ HW"),
    ("The Challenger Sale — Dixon & Adamson", "Book", "🤝 Sales"),
    ("SPIN Selling — Neil Rackham", "Book", "🤝 Sales"),
    ("MEDDIC Academy (meddic.com)", "Course", "🤝 Sales"),
    ("30 Minutes to President's Club (podcast)", "Podcast", "🤝 Sales"),
    ("Never Split the Difference — Chris Voss", "Book", "🤝 Sales"),
    ("ROS2 Documentation (docs.ros.org)", "Docs", "🤝 Sales"),
    ("Gong Revenue Intelligence Blog", "Blog", "🤝 Sales"),
    ("The Pyramid Principle — Barbara Minto", "Book", "🗣️ Comm"),
    ("On Writing Well — William Zinsser", "Book", "🗣️ Comm"),
    ("The Elements of Style — Strunk & White", "Book", "🗣️ Comm"),
    ("Smart Brevity — Axios / Springer", "Book", "🗣️ Comm"),
    ("Hemingway App (hemingwayapp.com)", "Tool", "🗣️ Comm"),
    ("Crucial Conversations — Patterson et al.", "Book", "🗣️ Comm"),
    ("Roger Love — 'Set Your Voice Free'", "Book/Audio", "🎙️ Voice"),
    ("Julian Treasure — TED Talk on speaking", "Video", "🎙️ Voice"),
    ("Toastmasters International", "Community", "🎙️ Voice"),
    ("Otter.ai (transcription)", "Tool", "🎙️ Voice"),
    ("Voice Memos (daily recordings)", "Tool", "🎙️ Voice"),
]

pillar_fill = {
    "🧠 LLM": LBLUE,
    "⚡ HW": "E0F5F5",
    "🤝 Sales": "FFF3E0",
    "🗣️ Comm": "F3E5F5",
    "🎙️ Voice": "FBE9E7",
}

for ri, (title, rtype, pillar) in enumerate(all_resources):
    row = 5 + ri
    bg = pillar_fill.get(pillar, WHITE)
    cell(rws, row, 2, title, bg=bg, fg="1A1A2E", sz=10, wrap=True)
    cell(rws, row, 3, rtype, bg=bg, fg=DGREY, sz=10, align_h="center")
    cell(rws, row, 4, pillar, bg=bg, fg=DGREY, sz=10, align_h="center")
    cell(rws, row, 5, "⬜ Not Started", bg=bg, fg=DGREY, sz=10, align_h="center")
    row_h(rws, row, 18)

# ── Reorder sheets ────────────────────────────────────────────
sheet_order = ["📊 Dashboard", "📅 Weekly Log", "🧠 LLM & RLHF", "⚡ Inference HW",
               "🤝 Tech Sales", "🗣️ Communication", "🎙️ Voice", "📚 Resources"]
for i, name in enumerate(sheet_order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

wb.save("/sessions/charming-confident-ride/mnt/outputs/Growth_Sprint_Tracker_2026.xlsx")
print("Done: Growth_Sprint_Tracker_2026.xlsx")
